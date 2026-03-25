import os
from pathlib import Path
import requests
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent.parent / ".env")

# Google Sheet: https://docs.google.com/spreadsheets/d/1iLYytIe_s_ZB5tUZLviEDijAItxE8-LlfrwH0ok1h60


def append_submission(name: str, email: str, description: str):
    webhook_url = os.getenv("GSHEET_WEBHOOK_URL", "")
    print(f"[Google Sheets] Attempting to save: name={name}, email={email}, desc={description[:50]}", flush=True)
    print(f"[Google Sheets] Webhook URL set: {bool(webhook_url)}", flush=True)
    print(f"[Google Sheets] Webhook URL value: {webhook_url[:30]}..." if webhook_url else "[Google Sheets] Webhook URL is EMPTY", flush=True)

    if webhook_url:
        try:
            response = requests.get(
                webhook_url,
                params={"name": name, "email": email, "description": description},
                timeout=15,
            )
            print(f"[Google Sheets] Response: {response.status_code} - {response.text[:100]}", flush=True)
            if response.status_code == 200 and "success" in response.text:
                print(f"[Google Sheets] Saved: {name}, {email}", flush=True)
                return
            else:
                print(f"[Google Sheets] Failed: HTTP {response.status_code}", flush=True)
        except Exception as e:
            print(f"[Google Sheets Error] {e}", flush=True)

    # Fallback to local Excel
    _save_local_excel(name, email, description)


def _save_local_excel(name: str, email: str, description: str):
    from openpyxl import Workbook, load_workbook

    file_path = os.getenv("SUBMISSION_PATH", "./submissions/user_submissions.xlsx")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(["Name", "Email_ID", "Requirement Description"])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

    ws.append([name, email, description])
    wb.save(file_path)
    print(f"[Local Excel Fallback] Saved: {name}, {email}", flush=True)
